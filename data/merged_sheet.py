from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


# ===== 設定 =====
BASE_DIR = Path("./out")        # out/sales pounds/data_YYYYMM.xlsx などがある場所
OUTPUT_DIR = Path("./merged")   # 生成先
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FILE_PATTERN = "data_*.xlsx"    # data_202505.xlsx 形式


def safe_sheet_name(name: str) -> str:
    """
    Excelシート名制約:
    - 31文字以内
    - \ / ? * [ ] : を含めない
    """
    name = name.strip()
    name = re.sub(r"[\\/?*\[\]:]", "_", name)
    return name[:31] if len(name) > 31 else name


def copy_sheet_values(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    """
    値をコピー（まずは安定重視：値のみ）
    ※ 罫線/色などのスタイルも必要なら拡張できます
    """
    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            dst_ws.cell(r, c).value = src_ws.cell(r, c).value


def parse_yyyymm_from_filename(path: Path) -> str:
    # data_202505.xlsx -> 202505
    m = re.match(r"data_(\d{6})\.xlsx$", path.name)
    if not m:
        raise ValueError(f"想定外のファイル名です: {path.name}")
    return m.group(1)


def find_month_files(base_dir: Path) -> Dict[str, List[Tuple[str, Path]]]:
    """
    戻り値:
      {
        "202505": [("sales pounds", out/sales pounds/data_202505.xlsx), ("level7", out/level7/data_202505.xlsx)],
        ...
      }
    """
    grouped: Dict[str, List[Tuple[str, Path]]] = {}

    for file_path in base_dir.glob(f"*/{FILE_PATTERN}"):
        if not file_path.is_file():
            continue
        yyyymm = parse_yyyymm_from_filename(file_path)
        folder_name = file_path.parent.name
        grouped.setdefault(yyyymm, []).append((folder_name, file_path))

    # 出力順を安定させる
    for yyyymm in grouped:
        grouped[yyyymm].sort(key=lambda x: x[0])  # folder_nameでソート

    return dict(sorted(grouped.items(), key=lambda x: x[0]))


def main():
    month_map = find_month_files(BASE_DIR)
    if not month_map:
        raise FileNotFoundError(
            f"{BASE_DIR}/<folder>/data_YYYYMM.xlsx が見つかりませんでした。")

    print("=== Merge start ===")

    for yyyymm, items in month_map.items():
        out_path = OUTPUT_DIR / f"data_{yyyymm}.xlsx"
        out_wb = Workbook()

        # 初期シート削除（空のSheetができるので）
        default_ws = out_wb.active
        out_wb.remove(default_ws)

        for folder_name, xlsx_path in items:
            src_wb = load_workbook(xlsx_path, data_only=True)

            # 各ファイルは「元シート名と同じシートが1枚」想定なので active を使う
            src_ws = src_wb.active

            sheet_name = safe_sheet_name(folder_name)
            dst_ws = out_wb.create_sheet(title=sheet_name)

            copy_sheet_values(src_ws, dst_ws)

        out_wb.save(out_path)
        print(f"created: {out_path}")

    print("=== Merge done ===")


if __name__ == "__main__":
    main()
