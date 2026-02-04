from __future__ import annotations

import re
import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ==========
# 設定
# ==========
INPUT_PATH = Path("data_b.xlsx")   # 例: Path("/mnt/data/data_b.xlsx")
OUTPUT_DIR = Path("./out")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ★追加：抽出したいシート名を指定（必須）
TARGET_SHEET_NAME = "level7"  # ←ここを対象シート名に変更

COL_PERIOD = "Period"
COL_CHANNEL = "channel_lv1"
QUARTER_CANDIDATES = ["Quater", "Quarter"]  # 表記ゆれ対応


def safe_folder_name(name: str) -> str:
    """
    フォルダ名に使えない文字を '_' に置換
    (Windows/Mac/Linux で安全にする)
    """
    name = name.strip()
    return re.sub(r'[\\/:*?"<>|\n\r\t]', "_", name)


# ==========
# ★追加：月名→月番号変換 & YYYYMM生成
# ==========
MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def normalize_year_month(year_value, month_value) -> str:
    """
    year_value: 2025 or 2025.0
    month_value: Aug / August / 8 / 08
    return: '202508'
    """
    year = int(year_value)

    # 月を数値に変換
    if isinstance(month_value, (int, float)):
        month = int(month_value)
    else:
        m = str(month_value).strip().lower()
        if m.isdigit():
            month = int(m)
        else:
            if m not in MONTH_MAP:
                raise ValueError(f"月の変換に失敗しました: {month_value}")
            month = MONTH_MAP[m]

    if not (1 <= month <= 12):
        raise ValueError(f"月が1〜12の範囲外です: {month_value} -> {month}")

    return f"{year}{month:02d}"


def find_header_row(ws: Worksheet) -> Tuple[int, int, str]:
    """
    同一行に Period / channel_lv1 / (Quater or Quarter) がある行をヘッダー行とみなす。
    戻り値: (header_row, channel_col, quarter_col_name)
    """
    for r in range(1, ws.max_row + 1):
        row_str: List[Tuple[int, str]] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                row_str.append((c, v.strip()))
        if not row_str:
            continue

        names = {name for _, name in row_str}
        if COL_PERIOD in names and COL_CHANNEL in names:
            qname = next((q for q in QUARTER_CANDIDATES if q in names), None)
            if qname:
                channel_col = next(
                    c for c, name in row_str if name == COL_CHANNEL)
                return r, channel_col, qname

    raise ValueError(
        f"ヘッダー行が見つかりません。'{COL_PERIOD}', '{COL_CHANNEL}', {QUARTER_CANDIDATES} のいずれかが同一行に必要です。"
    )


def read_headers_until_blank(ws: Worksheet, header_row: int, start_col: int) -> List[str]:
    """header_row の start_col から右へ、空セルまでをヘッダーとして読む"""
    headers: List[str] = []
    c = start_col
    while c <= ws.max_column:
        v = ws.cell(header_row, c).value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            break
        headers.append(str(v).strip())
        c += 1
    return headers


def read_rows(ws: Worksheet, header_row: int, start_col: int, width: int) -> List[List[Any]]:
    """header_row+1 から下へ、空行が続くまで width 分の行データを読む"""
    rows: List[List[Any]] = []
    r = header_row + 1
    empty_streak = 0

    while r <= ws.max_row:
        vals = [ws.cell(r, start_col + i).value for i in range(width)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            empty_streak += 1
            if empty_streak >= 3:
                break
        else:
            empty_streak = 0
            rows.append(vals)
        r += 1

    return rows


def write_month_file(
    out_path: Path,
    sheet_name: str,
    year_value: Any,
    quarter_value: Any,
    month_value: str,
    copy_headers: List[str],
    copy_rows: List[List[Any]],
):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ② ヘッダー情報
    ws["A2"].value = "Year"
    ws["A3"].value = "Quater"
    ws["A4"].value = "Period"
    ws["B2"].value = year_value
    ws["B3"].value = quarter_value
    ws["B4"].value = month_value

    # ③ データ本体（A7から、ヘッダー行含む）
    start_row, start_col = 7, 1  # A7
    for j, h in enumerate(copy_headers):
        ws.cell(start_row, start_col + j).value = h

    for i, row_vals in enumerate(copy_rows, start=1):
        for j, v in enumerate(row_vals):
            ws.cell(start_row + i, start_col + j).value = v

    wb.save(out_path)


def main():
    print("=== Split start ===")

    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"入力ファイルが見つかりません: {INPUT_PATH}")

    wb = load_workbook(INPUT_PATH, data_only=True)

    # ★追加：シート名指定で取得
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"指定したシート名が見つかりません: '{TARGET_SHEET_NAME}'. sheetnames={wb.sheetnames}")

    ws = wb[TARGET_SHEET_NAME]
    sheet_name = ws.title

    # ★追加：シート名フォルダに出力
    sheet_out_dir = OUTPUT_DIR / safe_folder_name(sheet_name)
    sheet_out_dir.mkdir(parents=True, exist_ok=True)

    # 年：B4
    year_value = ws["B4"].value
    if year_value is None:
        raise ValueError("B4 に年が入っていません。")

    # ヘッダー行を探す
    header_row, channel_col, quarter_col_name = find_header_row(ws)

    # 全列（Quarter, Period, channel_lv1, ...）を読み取る（グルーピング用）
    full_headers = read_headers_until_blank(ws, header_row, start_col=1)
    if not full_headers:
        raise ValueError("ヘッダーの読み取りに失敗しました（1列目から右に空白ばかり）。")

    full_rows = read_rows(ws, header_row, start_col=1, width=len(full_headers))

    # 列インデックス（分割用）
    header_map: Dict[str, int] = {h: i for i, h in enumerate(full_headers)}
    if COL_PERIOD not in header_map:
        raise ValueError(f"ヘッダーに '{COL_PERIOD}' がありません: {full_headers}")
    if quarter_col_name not in header_map:
        raise ValueError(f"ヘッダーに '{quarter_col_name}' がありません: {full_headers}")
    if COL_CHANNEL not in header_map:
        raise ValueError(f"ヘッダーに '{COL_CHANNEL}' がありません: {full_headers}")

    idx_period = header_map[COL_PERIOD]
    idx_quarter = header_map[quarter_col_name]
    idx_channel = header_map[COL_CHANNEL]

    # コピー対象は channel_lv1 以降（右側列）
    copy_headers = full_headers[idx_channel:]

    # 月ごとに分割（Period列でグルーピング）
    grouped: Dict[str, List[List[Any]]] = {}
    for row in full_rows:
        m = row[idx_period]
        if m is None or (isinstance(m, str) and m.strip() == ""):
            continue
        month = str(m).strip()
        grouped.setdefault(month, []).append(row)

    if not grouped:
        raise ValueError("Period列で分割できるデータが0件でした（Period列が空/別表記の可能性）。")

    # 月ごとに出力
    for month, month_full_rows in grouped.items():
        # Quarter は月内のユニーク値を連結
        q_vals = sorted({str(r[idx_quarter]).strip()
                        for r in month_full_rows if r[idx_quarter] is not None})
        quarter_value = ",".join(q_vals) if q_vals else None

        # コピー対象は右側列だけ切り出す
        month_copy_rows = [r[idx_channel:] for r in month_full_rows]

        # ファイル名：data_YYYYMM.xlsx
        yyyymm = normalize_year_month(year_value, month)
        out_path = sheet_out_dir / f"data_{yyyymm}.xlsx"

        write_month_file(
            out_path=out_path,
            sheet_name=sheet_name,
            year_value=int(year_value),   # 2025.0 → 2025
            quarter_value=quarter_value,
            month_value=month,
            copy_headers=copy_headers,
            copy_rows=month_copy_rows,
        )

        print(f"created: {out_path}")

    print("=== Split done ===")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("\n[ERROR]", f"{type(e).__name__}: {e}")
        traceback.print_exc()
        sys.exit(1)
